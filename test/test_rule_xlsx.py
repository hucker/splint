"""

The BaseCase.xlsx file contains:
  A           B
1 Status      Complete
2 Beginning   Yes
3 Middle	  Yes
4 Complete	  No
5 Post Mortem No
"""

import pytest
from openpyxl import load_workbook
import src.splint as splint


testdata = [


    # Testing cases for 'Sheet1' and specified columns
    ('Sheet1', 'A', 'B', 2, True, 'Beginning-Passed'),
    ('Sheet1', 'A', 'B', 3, True, 'Middle-Passed'),
    ('Sheet1', 'A', 'B', 4, False, 'Complete-Failed'),
    ('Sheet1', 'A', 'B', 5, False, 'Post Mortem-Failed'),

    # This gets the same pass/fail status, but there is no description
    # because using None for the desc-col has no default
    (None, None, None, 2, True, '-Passed'),
    (None, None, None, 3, True, '-Passed'),
    (None, None, None, 4, False, '-Failed'),
    (None, None, None, 5, False, '-Failed'),

]


@pytest.mark.parametrize('sheet, desc_col, val_col, row_start, expected_status, expected_msg', testdata)
def test_row_col_pass_fail_with_sheet(sheet, desc_col, val_col, row_start, expected_status, expected_msg):
    @splint.attributes(tag='foo')
    def check_pass_fail():
        wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
        yield from splint.rule_xlsx_a1_pass_fail(wb, sheet=sheet, val_col=val_col, row_start=str(row_start), desc_col=desc_col)

    s_func = splint.SplintFunction(check_pass_fail)
    ch = splint.SplintChecker(check_functions=[s_func], auto_setup=True)
    results = ch.run_all()

    # Verify that we got the status correct and that the messages are correct
    assert results[0].status is expected_status
    assert results[0].msg == expected_msg



@pytest.mark.parametrize('sheet, desc_col, val_col, row, expected_status, expected_msg', testdata)
def test_row_col_pass_fail_with_sheet_from_env(sheet, desc_col, val_col, row, expected_status, expected_msg):
    @splint.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from splint.rule_xlsx_a1_pass_fail(wb, sheet=sheet, val_col=val_col, row_start=str(row), desc_col=desc_col)

    # Test that we can load workbooks from the environment.  Same test as above.
    s_func = splint.SplintFunction(check_pass_fail)
    ch = splint.SplintChecker(check_functions=[s_func],
                              env={"wb": load_workbook('./rule_xlsx/BaseCase.xlsx')},
                              auto_setup=True)
    results = ch.run_all()

    # Verify that we got the status correct and that the messages are correct
    assert results[0].status is expected_status
    assert results[0].msg == expected_msg

def test_row_col_pass_fail_with_auto_detect():
    @splint.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from splint.rule_xlsx_a1_pass_fail(wb, sheet="Sheet1", val_col='B', row_start='2',row_end="auto", desc_col='A')

    s_func = splint.SplintFunction(check_pass_fail)
    wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
    ch = splint.SplintChecker(check_functions=[s_func],env={'wb':wb}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 4
    assert results[0].status is True
    assert results[1].status is True
    assert results[2].status is False
    assert results[3].status is False

def test_row_col_pass_fail_with_hardcoded():
    @splint.attributes(tag='foo')
    def check_pass_fail(wb):
        yield from splint.rule_xlsx_a1_pass_fail(wb, sheet="Sheet1", val_col='B', row_start='2',row_end="5", desc_col='A')

    s_func = splint.SplintFunction(check_pass_fail)
    wb = load_workbook('./rule_xlsx/BaseCase.xlsx')
    ch = splint.SplintChecker(check_functions=[s_func],env={'wb':wb}, auto_setup=True)
    results = ch.run_all()

    assert len(results) == 4
    assert results[0].status is True
    assert results[1].status is True
    assert results[2].status is False
    assert results[3].status is False
